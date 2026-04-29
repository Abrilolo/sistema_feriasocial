"""
Servicio para importar proyectos desde archivos Excel.
Realiza validación completa antes de crear proyectos en la BD.
"""
import uuid
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.project import Project
from app.models.user import User


class ExcelImportValidator:
    """Valida datos de un Excel antes de crear proyectos."""

    # Mapeo de nombres de columnas esperadas (case-insensitive)
    EXPECTED_COLUMNS = {
        "name": "name",
        "nombre": "name",
        "capacity": "capacity",
        "cupo": "capacity",
        "owner_email": "owner_email",
        "email_socio": "owner_email",
        "periodo": "periodo",
        "period": "periodo",
        "carreras_permitidas": "carreras_permitidas",
        "allowed_careers": "carreras_permitidas",
        "carreras": "carreras_permitidas",
        "objetivo": "objetivo",
        "objective": "objetivo",
        "actividades": "actividades",
        "activities": "actividades",
        "horario": "horario",
        "schedule": "horario",
        "competencias_requeridas": "competencias_requeridas",
        "required_skills": "competencias_requeridas",
        "competencias": "competencias_requeridas",
        "modalidad": "modalidad",
        "modality": "modalidad",
        "lugar_trabajo": "lugar_trabajo",
        "workplace": "lugar_trabajo",
        "lugar": "lugar_trabajo",
        "duracion": "duracion",
        "duration": "duracion",
        "poblacion_atendida": "poblacion_atendida",
        "served_population": "poblacion_atendida",
        "poblacion": "poblacion_atendida",
        "horas_acreditar": "horas_acreditar",
        "max_hours": "horas_acreditar",
        "horas": "horas_acreditar",
        "comentarios_adicionales": "comentarios_adicionales",
        "additional_comments": "comentarios_adicionales",
        "comentarios": "comentarios_adicionales",
        "clave_programa": "clave_programa",
        "program_key": "clave_programa",
        "clave": "clave_programa",
    }

    def __init__(self, file_content: bytes, owner_user_id: str | None = None):
        self.file_content = file_content
        self.fallback_owner_id = owner_user_id
        self.errors = []
        self.projects_data = []
        self.workbook = None
        self.worksheet = None

    def validate(self, db: Session) -> tuple[bool, list[dict], list[str]]:
        """
        Valida el Excel completo.
        Retorna (is_valid, projects_data, errors)
        """
        try:
            self._load_workbook()
            self._validate_structure()

            if self.errors:
                return False, [], self.errors

            self._extract_rows(db)

            if self.errors:
                return False, [], self.errors

            return True, self.projects_data, []

        except Exception as e:
            return False, [], [f"Error al procesar Excel: {str(e)}"]

    def _load_workbook(self):
        """Carga el workbook desde el contenido binario."""
        try:
            self.workbook = load_workbook(BytesIO(self.file_content), data_only=True)
            self.worksheet = self.workbook.active
        except Exception as e:
            self.errors.append(f"No se pudo leer el archivo Excel: {str(e)}")
            raise

    def _validate_structure(self):
        """Valida que el Excel tenga las columnas esperadas."""
        if not self.worksheet or self.worksheet.max_row == 0:
            self.errors.append("El archivo Excel está vacío.")
            return

        # Leer encabezados
        headers = []
        for cell in self.worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())

        if not headers:
            self.errors.append("No se encontraron encabezados en la primera fila.")
            return

        # Validar que al menos 'name' y 'capacity' existan
        has_name = any(h in self.EXPECTED_COLUMNS for h in headers)
        has_capacity = any(
            h in self.EXPECTED_COLUMNS and self.EXPECTED_COLUMNS[h] == "capacity"
            for h in headers
        )

        if not has_name:
            self.errors.append("Falta columna 'name' (nombre del proyecto) en el Excel.")
        if not has_capacity:
            self.errors.append("Falta columna 'capacity' (cupo) en el Excel.")

    def _extract_rows(self, db: Session):
        """Extrae y valida todas las filas del Excel."""
        # Mapear headers a índices
        headers = {}
        for i, cell in enumerate(self.worksheet[1], 1):
            if cell.value:
                col_name = str(cell.value).strip().lower()
                if col_name in self.EXPECTED_COLUMNS:
                    headers[self.EXPECTED_COLUMNS[col_name]] = i

        if not headers:
            self.errors.append("No se pudieron mapear las columnas.")
            return

        # Validar límite de filas
        if self.worksheet.max_row > 1001:  # 1 header + 1000 data
            self.errors.append("El archivo Excel tiene más de 1000 proyectos (límite excedido).")
            return

        # Procesar cada fila
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2), start=2):
            row_data = self._extract_row_data(row, headers, row_idx, db)

            if "error" in row_data:
                self.errors.append(row_data["error"])
            else:
                self.projects_data.append(row_data)

    def _extract_row_data(
        self, row, headers: dict, row_idx: int, db: Session
    ) -> dict:
        """Extrae y valida una fila individual."""
        # Extraer valores
        name = self._get_cell_value(row, headers.get("name"))
        capacity = self._get_cell_value(row, headers.get("capacity"))
        owner_email = self._get_cell_value(row, headers.get("owner_email"))

        # Validar campos obligatorios
        if not name or name.strip() == "":
            return {"error": f"Fila {row_idx}: 'name' está vacío."}

        name = name.strip()

        # Validar capacity
        try:
            if capacity is None or capacity == "":
                return {"error": f"Fila {row_idx}: 'capacity' está vacío."}
            capacity = int(capacity)
            if capacity <= 0:
                return {"error": f"Fila {row_idx}: 'capacity' debe ser mayor a 0."}
        except (ValueError, TypeError):
            return {
                "error": f"Fila {row_idx}: 'capacity' no es un número válido."
            }

        # Resolver owner_user_id
        owner_user_id = None

        # 1. Si hay owner_email en la fila, usarlo
        if owner_email and owner_email.strip():
            owner_email = owner_email.strip().lower()
            owner = db.query(User).filter(User.email == owner_email).first()
            if not owner:
                return {
                    "error": f"Fila {row_idx}: No existe usuario con email '{owner_email}'."
                }
            owner_user_id = str(owner.id)
        # 2. Si no, usar fallback
        elif self.fallback_owner_id:
            owner_user_id = self.fallback_owner_id
        else:
            return {
                "error": f"Fila {row_idx}: Sin 'owner_email' y sin owner por defecto."
            }

        # Validar que no haya duplicado (name + owner_id)
        existing = (
            db.query(Project)
            .filter(
                Project.name == name,
                Project.owner_user_id == owner_user_id,
            )
            .first()
        )
        if existing:
            return {
                "error": f"Fila {row_idx}: Ya existe proyecto '{name}' con este owner."
            }

        # Extraer campos opcionales
        data = {
            "id": str(uuid.uuid4()),
            "name": name,
            "capacity": capacity,
            "owner_user_id": owner_user_id,
            "description": self._get_cell_value(
                row, headers.get("description")
            ) or None,
            "periodo": self._get_cell_value(row, headers.get("periodo")) or None,
            "carreras_permitidas": self._get_cell_value(
                row, headers.get("carreras_permitidas")
            )
            or None,
            "objetivo": self._get_cell_value(row, headers.get("objetivo")) or None,
            "actividades": self._get_cell_value(row, headers.get("actividades"))
            or None,
            "horario": self._get_cell_value(row, headers.get("horario")) or None,
            "competencias_requeridas": self._get_cell_value(
                row, headers.get("competencias_requeridas")
            )
            or None,
            "modalidad": self._get_cell_value(row, headers.get("modalidad"))
            or None,
            "lugar_trabajo": self._get_cell_value(row, headers.get("lugar_trabajo"))
            or None,
            "duracion": self._get_cell_value(row, headers.get("duracion")) or None,
            "poblacion_atendida": self._get_cell_value(
                row, headers.get("poblacion_atendida")
            )
            or None,
            "horas_acreditar": self._get_cell_value(row, headers.get("horas_acreditar"))
            or None,
            "comentarios_adicionales": self._get_cell_value(
                row, headers.get("comentarios_adicionales")
            )
            or None,
            "clave_programa": self._get_cell_value(row, headers.get("clave_programa"))
            or None,
        }

        return data

    @staticmethod
    def _get_cell_value(row, col_idx: int | None) -> str | None:
        """Obtiene el valor de una celda, retorna None si está vacía o col_idx es None."""
        if col_idx is None:
            return None
        try:
            cell = row[col_idx - 1]  # openpyxl usa 0-based indexing
            value = cell.value
            if value is None or value == "":
                return None
            return str(value).strip()
        except (IndexError, AttributeError):
            return None
